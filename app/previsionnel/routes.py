from __future__ import annotations

from datetime import date
from io import BytesIO

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.extensions import db
from app.rbac import require_perm, can
from app.secteurs import get_secteur_labels
from app.models import (
    AppelProjetBudget,
    AppelProjetBudgetLigne,
    BudgetPrevisionnel,
    BudgetPrevisionnelLigne,
    Projet,
    Subvention,
)

bp = Blueprint("previsionnel", __name__, url_prefix="/previsionnel")


def _is_all_scope() -> bool:
    return bool(can("scope:all_secteurs"))


def _allowed_secteurs() -> list[str]:
    if _is_all_scope():
        return get_secteur_labels()
    secteur = (getattr(current_user, "secteur_assigne", None) or "").strip()
    return [secteur] if secteur else []


def _check_budget_scope(budget: BudgetPrevisionnel):
    if _is_all_scope():
        return
    if budget.secteur != (getattr(current_user, "secteur_assigne", None) or ""):
        abort(403)


def _can_edit() -> bool:
    return can("subventions:edit") or can("subventions:create") or can("depenses:create")


def _parse_float(value, default=0.0) -> float:
    raw = str(value or "").replace(" ", "").replace(",", ".")
    try:
        return float(raw) if raw else float(default)
    except Exception:
        return float(default)


def _parse_int(value, default=0) -> int:
    try:
        return int(value or default)
    except Exception:
        return int(default)


def _money(value) -> float:
    return round(float(value or 0), 2)


def _projects_for_budget(budget: BudgetPrevisionnel):
    return Projet.query.filter_by(secteur=budget.secteur).order_by(Projet.nom.asc()).all()


def _subventions_for_budget(budget: BudgetPrevisionnel):
    return (
        Subvention.query.filter_by(secteur=budget.secteur, annee_exercice=budget.annee)
        .order_by(Subvention.nom.asc())
        .all()
    )


def _project_totals(lines):
    totals = {}
    for line in lines:
        key = line.projet_id or 0
        label = line.projet.nom if line.projet else "Transversal / non affecté"
        bucket = totals.setdefault(key, {"label": label, "charges": 0.0, "produits": 0.0})
        if line.nature == "produit":
            bucket["produits"] += float(line.montant or 0)
        else:
            bucket["charges"] += float(line.montant or 0)
    rows = []
    for data in totals.values():
        rows.append({
            "label": data["label"],
            "charges": _money(data["charges"]),
            "produits": _money(data["produits"]),
            "solde": _money(data["produits"] - data["charges"]),
        })
    return sorted(rows, key=lambda r: r["label"].lower())


def _add_table(ws, name: str, start_row: int, end_row: int, end_col: int, style="TableStyleMedium9"):
    if end_row <= start_row:
        return
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    safe = "".join(ch if ch.isalnum() else "_" for ch in name)
    if not safe or not safe[0].isalpha():
        safe = f"T_{safe or 'Table'}"
    existing = set(ws.parent.worksheets[0].tables.keys()) if ws.parent.worksheets else set()
    for sh in ws.parent.worksheets:
        existing.update(sh.tables.keys())
    base = safe[:180]
    idx = 1
    while safe in existing:
        idx += 1
        safe = f"{base}_{idx}"
    table = Table(displayName=safe, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _style_ws(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00 €'
    ws.sheet_view.showGridLines = False
    for col in range(1, ws.max_column + 1):
        width = 12
        for cell in ws[get_column_letter(col)]:
            width = max(width, min(38, len(str(cell.value or "")) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_budget_workbook(budget: BudgetPrevisionnel, lines):
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget prévisionnel"
    ws.append([f"Budget prévisionnel — {budget.nom}"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append(["Année", budget.annee, "Secteur", budget.secteur, "Statut", budget.statut])
    ws.append([])
    ws.append(["Synthèse", "Montant"])
    ws.append(["Charges prévisionnelles", budget.total_charges])
    ws.append(["Produits prévisionnels", budget.total_produits])
    ws.append(["Solde produits - charges", budget.solde])
    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["Nature", "Compte", "Libellé", "Projet", "Montant", "Commentaire"])
    for line in lines:
        ws.append([
            "Produit" if line.nature == "produit" else "Charge",
            line.compte,
            line.libelle,
            line.projet.nom if line.projet else "Transversal / non affecté",
            float(line.montant or 0),
            line.commentaire or "",
        ])
    _add_table(ws, "BudgetPrevisionnelLignes", header_row, ws.max_row, 6)
    ws.freeze_panes = f"A{header_row + 1}"

    ws2 = wb.create_sheet("Par projet")
    ws2.append(["Projet", "Charges", "Produits", "Solde"])
    for row in _project_totals(lines):
        ws2.append([row["label"], row["charges"], row["produits"], row["solde"]])
    _add_table(ws2, "BudgetPrevisionnelParProjet", 1, ws2.max_row, 4, style="TableStyleMedium4")
    ws2.freeze_panes = "A2"

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(1, col).font = Font(bold=True)
        _style_ws(sheet)
    return wb


def _build_appel_workbook(appel: AppelProjetBudget):
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget appel à projet"
    ws.append([f"Budget appel à projet — {appel.nom}"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append(["Budget source", appel.budget.nom, "Projet", appel.projet.nom if appel.projet else "", "Financeur", appel.financeur or ""])
    ws.append([])
    ws.append(["Synthèse", "Montant"])
    ws.append(["Charges retenues", appel.total_charges])
    ws.append(["Produits retenus", appel.total_produits])
    ws.append(["Solde produits - charges", appel.solde])
    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["Nature", "Compte", "Libellé", "Projet source", "Montant source", "Montant retenu", "% retenu", "Commentaire"])
    for row in appel.lignes:
        line = row.ligne_budget
        ws.append([
            "Produit" if line.nature == "produit" else "Charge",
            line.compte,
            line.libelle,
            line.projet.nom if line.projet else "Transversal / non affecté",
            float(line.montant or 0),
            float(row.montant_retenu or 0),
            float(row.pourcentage_retenu or 0) if row.pourcentage_retenu is not None else "",
            row.commentaire or "",
        ])
    _add_table(ws, "BudgetAppelProjet", header_row, ws.max_row, 8)
    ws.freeze_panes = f"A{header_row + 1}"
    _style_ws(ws)
    return wb


@bp.route("/", methods=["GET", "POST"])
@login_required
@require_perm("subventions:view")
def index():
    secteurs = _allowed_secteurs()
    if request.method == "POST":
        if not _can_edit():
            abort(403)
        annee = _parse_int(request.form.get("annee"), date.today().year)
        secteur = (request.form.get("secteur") or "").strip()
        nom = (request.form.get("nom") or f"Budget prévisionnel {secteur} {annee}").strip()
        if secteur not in secteurs:
            abort(403)
        budget = BudgetPrevisionnel(nom=nom, annee=annee, secteur=secteur, statut="brouillon")
        db.session.add(budget)
        db.session.commit()
        flash("Budget prévisionnel créé.", "success")
        return redirect(url_for("previsionnel.detail", budget_id=budget.id))

    q = BudgetPrevisionnel.query
    if not _is_all_scope():
        q = q.filter(BudgetPrevisionnel.secteur.in_(secteurs or ["__none__"]))
    budgets = q.order_by(BudgetPrevisionnel.annee.desc(), BudgetPrevisionnel.secteur.asc(), BudgetPrevisionnel.nom.asc()).all()
    return render_template("previsionnel/index.html", budgets=budgets, secteurs=secteurs, current_year=date.today().year, can_edit_budget=_can_edit())


@bp.route("/<int:budget_id>", methods=["GET", "POST"])
@login_required
@require_perm("subventions:view")
def detail(budget_id: int):
    budget = BudgetPrevisionnel.query.get_or_404(budget_id)
    _check_budget_scope(budget)

    if request.method == "POST":
        if not _can_edit():
            abort(403)
        action = request.form.get("action") or ""

        if action == "update_budget":
            budget.nom = (request.form.get("nom") or budget.nom).strip()
            budget.statut = (request.form.get("statut") or budget.statut or "brouillon").strip()
            budget.notes = (request.form.get("notes") or "").strip() or None
            db.session.commit()
            flash("Budget mis à jour.", "success")
            return redirect(url_for("previsionnel.detail", budget_id=budget.id))

        if action == "add_line":
            nature = (request.form.get("nature") or "charge").strip()
            if nature not in ("charge", "produit"):
                nature = "charge"
            projet_id = _parse_int(request.form.get("projet_id"), 0) or None
            if projet_id:
                projet = Projet.query.get_or_404(projet_id)
                if projet.secteur != budget.secteur:
                    abort(400)
            line = BudgetPrevisionnelLigne(
                budget_id=budget.id,
                nature=nature,
                compte=(request.form.get("compte") or ("74" if nature == "produit" else "60")).strip(),
                libelle=(request.form.get("libelle") or "").strip(),
                montant=_parse_float(request.form.get("montant")),
                projet_id=projet_id,
                commentaire=(request.form.get("commentaire") or "").strip() or None,
                ordre=len(budget.lignes) + 1,
            )
            if not line.libelle:
                flash("Le libellé de la ligne est obligatoire.", "danger")
                return redirect(url_for("previsionnel.detail", budget_id=budget.id))
            db.session.add(line)
            db.session.commit()
            flash("Ligne ajoutée au prévisionnel.", "success")
            return redirect(url_for("previsionnel.detail", budget_id=budget.id))

        if action == "delete_line":
            line = BudgetPrevisionnelLigne.query.get_or_404(_parse_int(request.form.get("line_id")))
            if line.budget_id != budget.id:
                abort(400)
            if line.lignes_appel:
                flash("Cette ligne est déjà utilisée dans un budget d'appel à projet : suppression bloquée.", "danger")
                return redirect(url_for("previsionnel.detail", budget_id=budget.id))
            db.session.delete(line)
            db.session.commit()
            flash("Ligne supprimée.", "warning")
            return redirect(url_for("previsionnel.detail", budget_id=budget.id))

        if action == "create_appel":
            nom = (request.form.get("appel_nom") or "").strip() or f"Budget appel à projet — {budget.nom}"
            projet_id = _parse_int(request.form.get("appel_projet_id"), 0) or None
            subvention_id = _parse_int(request.form.get("subvention_id"), 0) or None
            financeur = (request.form.get("financeur") or "").strip() or None
            if subvention_id:
                sub = Subvention.query.get_or_404(subvention_id)
                if sub.secteur != budget.secteur:
                    abort(400)
                financeur = financeur or sub.nom
            if projet_id:
                projet = Projet.query.get_or_404(projet_id)
                if projet.secteur != budget.secteur:
                    abort(400)
            selected_ids = [int(x) for x in request.form.getlist("include_line") if str(x).isdigit()]
            if not selected_ids:
                flash("Sélectionne au moins une ligne à inclure dans le budget d'appel à projet.", "danger")
                return redirect(url_for("previsionnel.detail", budget_id=budget.id))
            appel = AppelProjetBudget(budget_id=budget.id, projet_id=projet_id, subvention_id=subvention_id, nom=nom, financeur=financeur)
            db.session.add(appel)
            db.session.flush()
            for line in BudgetPrevisionnelLigne.query.filter(BudgetPrevisionnelLigne.id.in_(selected_ids), BudgetPrevisionnelLigne.budget_id == budget.id).all():
                pct = _parse_float(request.form.get(f"pct_{line.id}"), 100.0)
                retained = round(float(line.montant or 0) * pct / 100.0, 2)
                manual = request.form.get(f"montant_{line.id}")
                if manual not in (None, ""):
                    retained = _parse_float(manual)
                db.session.add(AppelProjetBudgetLigne(appel_id=appel.id, budget_ligne_id=line.id, montant_retenu=retained, pourcentage_retenu=pct))
            db.session.commit()
            flash("Budget d'appel à projet généré depuis le prévisionnel.", "success")
            return redirect(url_for("previsionnel.appel_detail", appel_id=appel.id))

        abort(400)

    lines = BudgetPrevisionnelLigne.query.filter_by(budget_id=budget.id).order_by(BudgetPrevisionnelLigne.nature.asc(), BudgetPrevisionnelLigne.compte.asc(), BudgetPrevisionnelLigne.ordre.asc(), BudgetPrevisionnelLigne.id.asc()).all()
    projects = _projects_for_budget(budget)
    subventions = _subventions_for_budget(budget)
    appels = AppelProjetBudget.query.filter_by(budget_id=budget.id).order_by(AppelProjetBudget.created_at.desc()).all()
    project_rows = _project_totals(lines)
    return render_template("previsionnel/detail.html", budget=budget, lines=lines, projects=projects, subventions=subventions, appels=appels, project_rows=project_rows, can_edit_budget=_can_edit())


@bp.route("/<int:budget_id>/export.xlsx")
@login_required
@require_perm("subventions:view")
def export_budget(budget_id: int):
    budget = BudgetPrevisionnel.query.get_or_404(budget_id)
    _check_budget_scope(budget)
    lines = BudgetPrevisionnelLigne.query.filter_by(budget_id=budget.id).order_by(BudgetPrevisionnelLigne.nature.asc(), BudgetPrevisionnelLigne.compte.asc(), BudgetPrevisionnelLigne.id.asc()).all()
    wb = _build_budget_workbook(budget, lines)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"budget_previsionnel_{budget.annee}_{budget.secteur}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/appel/<int:appel_id>")
@login_required
@require_perm("subventions:view")
def appel_detail(appel_id: int):
    appel = AppelProjetBudget.query.get_or_404(appel_id)
    _check_budget_scope(appel.budget)
    return render_template("previsionnel/appel_detail.html", appel=appel, can_edit_budget=_can_edit())


@bp.route("/appel/<int:appel_id>/delete", methods=["POST"])
@login_required
@require_perm("subventions:view")
def appel_delete(appel_id: int):
    if not _can_edit():
        abort(403)
    appel = AppelProjetBudget.query.get_or_404(appel_id)
    _check_budget_scope(appel.budget)
    budget_id = appel.budget_id
    db.session.delete(appel)
    db.session.commit()
    flash("Budget d'appel à projet supprimé.", "warning")
    return redirect(url_for("previsionnel.detail", budget_id=budget_id))


@bp.route("/appel/<int:appel_id>/export.xlsx")
@login_required
@require_perm("subventions:view")
def export_appel(appel_id: int):
    appel = AppelProjetBudget.query.get_or_404(appel_id)
    _check_budget_scope(appel.budget)
    wb = _build_appel_workbook(appel)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"budget_appel_projet_{appel.id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
